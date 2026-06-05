from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

# =========================================================
# 1. 用 openpyxl 创建一个带合并单元格、公式的 Excel 文件
#    演示：工作簿(Workbook) → 工作表(Worksheet) → 单元格(Cell)
# =========================================================
wb = Workbook()                     # 创建工作簿对象
ws = wb.active                      # 获取活动工作表
ws.title = "销售明细"                # 修改工作表名

# 写入表头
headers = ["销售人员", "产品", "销售额", "奖金"]
ws.append(headers)                  # 追加整行

# 写入示例数据
data = [
    ("张三", "A产品", 12500, "=C2*0.1"),   # 奖金 = 销售额*0.1 (公式)
    ("李四", "B产品", 15800, "=C3*0.1"),
    ("王五", "A产品", 13200, "=C4*0.1"),
    ("赵六", "C产品", 17900, "=C5*0.1"),
]
for row_data in data:
    ws.append(row_data)

# 合并标题行上方的单元格（演示合并单元格）
ws.merge_cells("A1:D1")              # 合并 A1:D1
ws["A1"] = "2026年第一季度销售报表"    # 合并后左上角赋值
ws["A1"].font = Font(bold=True, size=14)

# 保存初始文件（先保存，后面再用 pandas 和 openpyxl 分别读取）
wb.save("销售数据.xlsx")
wb.close()
print("✅ 已创建文件：销售数据.xlsx（包含合并单元格、公式）\n")

# =========================================================
# 2. openpyxl 面向对象方式逐单元格处理
#    特点：可以读写公式、修改格式、处理合并单元格等底层细节。
# =========================================================
print("\n【Openpyxl 面向对象操作部分】")
# 2.1 以只读模式打开大文件（演示 read_only、data_only 参数）
wb_readonly = load_workbook(
    "销售数据.xlsx",
    read_only=True,      # 只读模式，无法保存，但速度快、内存小
    data_only=False      # 不取公式的值，取公式本身（为了演示公式读取）
)
ws_readonly = wb_readonly["销售明细"]

print(f"工作表名: {ws_readonly.title}")
print(f"最大行: {ws_readonly.max_row}，最大列: {ws_readonly.max_column}")

# 演示读取公式（因为 data_only=False）
cell_formula = ws_readonly["D3"].value   # D3 是李四的奖金单元格
print(f"D3 单元格原始内容（公式）: {cell_formula}")

# 关闭只读工作簿（必须 close 释放资源）
wb_readonly.close()
print("只读工作簿已关闭。")

#2.2 用读写模式打开，修改单元格样式、取消合并单元格
wb_rw = load_workbook("销售数据.xlsx")
ws_rw = wb_rw["销售明细"]

# 处理合并单元格 – 获取合并区域并取消
print(f"\n合并单元格区域: {list(ws_rw.merged_cells.ranges)}")  # merged_cells.ranges
ws_rw.unmerge_cells("A1:D1")          # 取消合并

# 遍历行，为“高销售额”标红加粗
for row in ws_rw.iter_rows(min_row=2, max_col=4, values_only=False):
    sales_cell = row[2]   # 第3列（销售额）
    if sales_cell.value and sales_cell.value > 15000:
        name_cell = row[0]
        name_cell.font = Font(color="FF0000", bold=True)   # 红色加粗

# 保存新文件
new_file = "销售数据_修改后.xlsx"
wb_rw.save(new_file)
wb_rw.close()
print(f"✅ 已生成带格式和评级的文件：{new_file}")




import pandas as pd
# ========== 1. 写入示例数据（pandas 负责生成） ==========
df_original = pd.DataFrame({
    "销售人员": ["张三", "李四", "王五", "赵六", "孙七"],
    "部门": ["销售部", "技术部", "销售部", "技术部", "销售部"],
    "销售额": [12500, 15800, 13200, 17900, 14600],
    "奖金": [1250, 1580, 1320, 1790, 1460]
})
df_original.to_excel("绩效数据.xlsx", sheet_name="一月绩效", index=False)
print("✅ 用 pandas 生成示例 Excel 文件\n")

# ========== 2. Pandas 数据分析 -> 展示常用方法 ==========
print("【Pandas 核心方法演示】")
# 读取指定工作表
df = pd.read_excel("绩效数据.xlsx", sheet_name="一月绩效")

# 基础信息
print("\n1. df.head(2) ：前2行数据")
print(df.head(2))

print("\n2. df.shape ：数据形状 (行数, 列数)")
print(df.shape)

print("\n3. df.columns.tolist() ：列名列表")
print(df.columns.tolist())

print("\n4. df.info() ：数据类型与非空统计")
df.info()

print("\n5. df.describe() ：数值列统计摘要")
print(df.describe())

# 数据选择与切片
print("\n6. df['销售额'] ：选择单列")
print(df["销售额"].head(2))

print("\n7. df[['销售人员','销售额']] ：选择多列")
print(df[["销售人员", "销售额"]].head(2))

print("\n8. df.iloc[:3, :2] ：按位置切片 (前3行,前2列)")
print(df.iloc[:3, :2])

# 条件筛选
print("\n9. df[df['销售额'] > 15000] ：筛选销售额 > 15000")
print(df[df["销售额"] > 15000])

# 缺失值处理 (模拟一个空值)
df.loc[2, "奖金"] = None
print("\n10. 原始有缺失:\n", df[["销售人员","奖金"]].head(3))
df["奖金"] = df["奖金"].fillna(0)   # fillna() 填充缺失
print("    fillna(0) 后:\n", df[["销售人员","奖金"]].head(3))

# 分组聚合
print("\n11. groupby('部门')['销售额'].sum()  -> 部门销售总额")
grouped = df.groupby("部门")["销售额"].sum()
print(grouped)

# 多列聚合
agg_result = df.groupby("部门").agg(
    总销售额=("销售额", "sum"),
    平均奖金=("奖金", "mean")
)
print("\n12. agg() 多列聚合:\n", agg_result)

# 排序
df_sorted = df.sort_values("销售额", ascending=False)
print("\n13. sort_values 降序（销售额）:\n", df_sorted[["销售人员","销售额"]])

# 输出统计后的结果至新 Excel
grouped.to_excel("部门统计.xlsx", header=True)
print("\n14. 已将聚合结果写入 '部门统计.xlsx'")