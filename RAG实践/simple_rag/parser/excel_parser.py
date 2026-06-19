from typing import List
import os


class ExcelParser:
    """Excel 文档解析器"""

    def parse(self, file_path: str) -> List[str]:
        try:
            import pandas as pd
            texts = []
            excel_file = pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                texts.append(f"【工作表: {sheet_name}】")
                headers = df.columns.tolist()
                texts.append(f"列名: {', '.join(str(h) for h in headers)}")
                for idx, row in df.iterrows():
                    row_text = " | ".join(f"{h}: {v}" for h, v in zip(headers, row.values) if pd.notna(v))
                    if row_text:
                        texts.append(row_text)
                texts.append("")
            return texts
        except ImportError:
            try:
                from openpyxl import load_workbook
                texts = []
                wb = load_workbook(file_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    texts.append(f"【工作表: {sheet_name}】")
                    rows_data = []
                    for row in ws.iter_rows(values_only=True):
                        row_text = " | ".join(f"{v}" for v in row if v is not None)
                        if row_text:
                            rows_data.append(row_text)
                    texts.append("表头: " + rows_data[0] if rows_data else "")
                    texts.extend(rows_data[1:] if len(rows_data) > 1 else [])
                    texts.append("")
                wb.close()
                return texts
            except Exception as e:
                print(f"Excel 解析失败: {e}")
                return []


if __name__ == "__main__":
    parser = ExcelParser()
    # result = parser.parse("example.xlsx")
    # print(result)
